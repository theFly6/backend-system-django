from django.http import HttpResponse
from django.shortcuts import render, redirect

from app01.models import Department


def depart_lst(request):
    """部门列表"""
    data_lst = Department.objects.all()
    return render(request, 'depart_list.html', {'data_lst': data_lst})


def depart_add(request):
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    title = request.POST.get('title')
    Department.objects.create(title=title)
    return redirect('/depart/list')


def depart_del(request):
    depart_id = request.GET.get('uid')
    Department.objects.filter(id=depart_id).delete()
    return redirect('/depart/list')


def depart_edit(request, depart_id):
    depart = Department.objects.filter(id=depart_id)
    if request.method == 'GET':
        return render(request, 'depart_edit.html', {'edit_title': depart.first().title})
    depart.update(title=request.POST.get('title'))
    return redirect('/depart/list')


def depart_upload(request):
    # from django.core.files.uploadedfile import InMemoryUploadedFile
    from openpyxl import load_workbook
    # 1.获取文件对象
    exc_file = request.FILES.get('exc_file')
    # 2.通过load_workbook打开对象
    wb = load_workbook(exc_file)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        title = row[0].value
        if not Department.objects.filter(title=title).exists():
            Department.objects.create(title=title)
    return redirect('/depart/list')
